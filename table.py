import sys
from PyQt5.QtWidgets import QTableWidget, QApplication, QMainWindow, QTableWidgetItem,QHeaderView,QSpinBox,QWidget
from PyQt5.QtCore import Qt
# import load_workbook
from openpyxl import load_workbook,Workbook
import pandas as pd



class MyTable(QTableWidget):
    count=0
    list_Changed_Cells=[]
    list_Cells=[]
    r=0
    c=0
    def __init__(self, r, c):
        super().__init__(r, c)
        self.init_ui()
        self.r=r
        self.c=c


        header = self.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)




    def init_ui(self):
        self.cellChanged.connect(self.c_current)
        self.show()

    def c_current(self):
        self.count=self.count+1

        row = self.currentRow()
        col = self.currentColumn()
        data = self.item(row, col)
        #data = data.text()

        self.list_Cells.append([row+1,col+1,data.text()])


        if self.count >Sheet.count:
            self.list_Changed_Cells.append([row+1,col+1,data.text()])





#without Sheet name

class Sheet(QMainWindow):
    list_with_values = []
    form_widget=MyTable
    count = 0



    def __init__(self,path,sheetName):
        super().__init__()
        filepath = path
        wb = load_workbook(filepath,read_only=True)
        if sheetName is None:
            sheet = wb.active
        else:
            sheet=wb[sheetName]

        max_row = sheet.max_row
        max_column = sheet.max_column
        self.form_widget = MyTable(max_row, max_column)

        #self.form_widget.resizeColumnsToContents()
        self.setCentralWidget(self.form_widget)

        self.list_with_values.clear()
        for cell in sheet[1]:
            self.list_with_values.append(cell.value)
        self.form_widget.setHorizontalHeaderLabels(self.list_with_values)
        f=0
        for row in sheet:
          if f==0:
              f=1
          else:
            for cell in row:
                # get particular cell value
                cell_obj = cell#sheet.cell(row=i, column=j)
                if cell_obj.value is not None:
                    x=str(cell_obj.value)
                    number = QTableWidgetItem(x)
                    self.form_widget.setCurrentCell(cell_obj.row-1, cell_obj.column-1)
                    self.form_widget.setItem(cell_obj.row-1,cell_obj.column-1, number)
                    self.count=self.count+1
        self.show()
        wb.close()



    def SaveAs(self,path):
            wb=Workbook()

            ws = wb.active

            col=0
            for i in self.list_with_values:
                col=col+1
                ws.cell(row=1,column=col).value=i
            values=MyTable.list_Cells
            for i in values:
                ws.cell(row=i[0], column=i[1]).value = i[2]

            wb.save(path)
            wb.close()

    def setData(self, index, value, role=Qt.EditRole):
        if index.isValid():
            row = index.row()
            col = index.column()
            self._data.iloc[row][col] = value
            self.dataChanged.emit(index, index, (Qt.DisplayRole,))
            return True
        return False
    def setData(self, index, value, role):
        if not index.isValid():
            return False
        if role != QtCore.Qt.EditRole:
            return False
        row = index.row()
        if row < 0 or row >= len(self._data.values):
            return False
        column = index.column()
        if column < 0 or column >= self._data.columns.size:
            return False
        self._data.values[row][column] = value
        self.dataChanged.emit(index, index)
        return True
    def save(self,path,sheetName):
        wb= load_workbook(path)

        if sheetName is None:
            sheet = wb.active
        else:
            sheet = wb[sheetName]

        changed_values=MyTable.list_Changed_Cells

        for i in changed_values:
            sheet.cell(row=i[0], column=i[1]).value = i[2]
        wb.save(path)
        wb.close()


