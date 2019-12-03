from PyQt5 import QtCore, QtGui, QtWidgets
from tkinter import Tk,messagebox
from tkinter.filedialog import askopenfilename,asksaveasfilename
import os,sys
from PyQt5.QtWidgets import QTableWidget, QApplication, QMainWindow, QTableWidgetItem,QMessageBox,QTableView
from PyQt5 import QtCore, QtGui
from openpyxl import load_workbook
import table
import pandatest
from ntpath import basename
import pandas as pd
import numpy as np
Qt = QtCore.Qt
from pandas import ExcelWriter
from PyQt5 import QtCore, QtGui, QtWidgets
from pip._vendor.requests import head
from configparser import SafeConfigParser
import hqdriver

class Ui_Dialog_ComboBox(object):
    def setupUi(self, Dialog,sheetnames):
        Dialog.setObjectName("Dialog")
        Dialog.resize(259, 126)
        self.gridLayout = QtWidgets.QGridLayout(Dialog)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_3 = QtWidgets.QLabel(Dialog)
        self.label_3.setText("")
        self.label_3.setPixmap(QtGui.QPixmap("icons8-question-mark-64.png"))
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_3.addWidget(self.label_3)
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_3.addWidget(self.label_2)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem)
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        spacerItem1 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.horizontalLayout_4.addItem(spacerItem1)
        self.comboBox = QtWidgets.QComboBox(Dialog)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItems(sheetnames)
        self.comboBox.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.horizontalLayout_4.addWidget(self.comboBox)
        spacerItem2 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.horizontalLayout_4.addItem(spacerItem2)
        self.verticalLayout_2.addLayout(self.horizontalLayout_4)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem3)
        self.pushButton_ok = QtWidgets.QPushButton(Dialog)
        self.pushButton_ok.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_ok.setObjectName("pushButton_ok")
        self.horizontalLayout_2.addWidget(self.pushButton_ok)
        self.pushButton_Cancel = QtWidgets.QPushButton(Dialog)
        self.pushButton_Cancel.setObjectName("pushButton_Cancel")
        self.pushButton_Cancel.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.horizontalLayout_2.addWidget(self.pushButton_Cancel)
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem4)
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.gridLayout.addLayout(self.verticalLayout_2, 0, 0, 1, 1)
        self.pushButton_ok.clicked.connect(Dialog.reject)
        self.pushButton_Cancel.clicked.connect(self.on_click_cancel)



        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        Dialog.exec_()



    def on_click_cancel(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)

        msg.setText("Please Select a sheet")

        msg.setWindowTitle("Warning")

        msg.setStandardButtons(QMessageBox.Ok)


        msg.exec_()

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Select Worksheet"))
        self.label_2.setText(_translate("Dialog", "Which Worksheet do you want to import?"))
        self.pushButton_ok.setText(_translate("Dialog", "ok"))
        self.pushButton_Cancel.setText(_translate("Dialog", "cancel"))

class Ui_MainWindow(object):
    parser = SafeConfigParser()
    parser.read('C:/Users/LIJAMANNATHARA/PycharmProjects/UIAF_UI/test  output/config.properties')
    path = os.getcwd()
    masterSelectedFile = None
    testScriptSelectedFile=None
    testComponentSelectedFile=None
    testObjectsSelectedFile=None
    testDataSelectedFile=None
    ExecutionFile=None
    table_widget = table.Sheet
    script_sheet=None
    Component_Sheet=None
    Object_sheet=None
    Data_sheet=None
    Execution_sheet=None
    script_view = None
    script_model=None
    component_view=None
    component_model=None
    set_view=None
    set_model=None
    data_view=None
    data_model=None
    object_view=None
    object_model=None
    execution_view=None
    execution_model=None
    your_pandas_data_scripts=None
    your_pandas_data_components=None
    your_pandas_data_sets = None
    your_pandas_data_data = None
    your_pandas_data_objects = None
    your_pandas_data_execution = None


    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1900, 800)
        MainWindow.setMaximumSize(QtCore.QSize(1900, 800))
        MainWindow.setStyleSheet("")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.Maintab = QtWidgets.QTabWidget(self.centralwidget)
        self.Maintab.setStyleSheet("font: 75 10pt \"Arial\";")
        self.Maintab.setObjectName("Maintab")
        self.RecordTab = QtWidgets.QWidget()
        self.RecordTab.setObjectName("RecordTab")
        self.gridLayout_10 = QtWidgets.QGridLayout(self.RecordTab)
        self.gridLayout_10.setObjectName("gridLayout_10")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.RecordPlayBtn = QtWidgets.QPushButton(self.RecordTab)
        self.RecordPlayBtn.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.RecordPlayBtn.setObjectName("RecordPlayBtn")
        self.verticalLayout.addWidget(self.RecordPlayBtn, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.verticalLayout_32 = QtWidgets.QVBoxLayout()
        self.verticalLayout_32.setObjectName("verticalLayout_32")
        self.textBrowser_RecordPlay = QtWidgets.QTextBrowser(self.RecordTab)
        self.textBrowser_RecordPlay.setObjectName("textBrowser_RecordPlay")
        self.verticalLayout_32.addWidget(self.textBrowser_RecordPlay)
        self.label_17 = QtWidgets.QLabel(self.RecordTab)
        self.label_17.setObjectName("label_17")
        self.verticalLayout_32.addWidget(self.label_17, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.verticalLayout.addLayout(self.verticalLayout_32)
        self.gridLayout_10.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.Maintab.addTab(self.RecordTab, "")
        self.ScriptGenerateTab = QtWidgets.QWidget()
        self.ScriptGenerateTab.setObjectName("ScriptGenerateTab")
        self.gridLayout_11 = QtWidgets.QGridLayout(self.ScriptGenerateTab)
        self.gridLayout_11.setObjectName("gridLayout_11")
        self.verticalLayout_16 = QtWidgets.QVBoxLayout()
        self.verticalLayout_16.setObjectName("verticalLayout_16")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem)
        self.label_WorkBookName_ScriptGenerate = QtWidgets.QLabel(self.ScriptGenerateTab)
        self.label_WorkBookName_ScriptGenerate.setObjectName("label_WorkBookName_ScriptGenerate")
        self.horizontalLayout_4.addWidget(self.label_WorkBookName_ScriptGenerate, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.label_WorkBookName2_ScriptGenerate = QtWidgets.QLabel(self.ScriptGenerateTab)
        self.label_WorkBookName2_ScriptGenerate.setObjectName("label_WorkBookName2_ScriptGenerate")
        self.horizontalLayout_4.addWidget(self.label_WorkBookName2_ScriptGenerate, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.pushButton_Open_ScriptGenerate = QtWidgets.QPushButton(self.ScriptGenerateTab)
        self.pushButton_Open_ScriptGenerate.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_ScriptGenerate.setObjectName("pushButton_Open_ScriptGenerate")
        self.horizontalLayout_4.addWidget(self.pushButton_Open_ScriptGenerate, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.pushButton_Save_ScriptGenerate = QtWidgets.QPushButton(self.ScriptGenerateTab)
        self.pushButton_Save_ScriptGenerate.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_ScriptGenerate.setObjectName("pushButton_Save_ScriptGenerate")
        self.horizontalLayout_4.addWidget(self.pushButton_Save_ScriptGenerate, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.pushButton_GenerateScript_ScriptGenerate = QtWidgets.QPushButton(self.ScriptGenerateTab)
        self.pushButton_GenerateScript_ScriptGenerate.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_GenerateScript_ScriptGenerate.setObjectName("pushButton_GenerateScript_ScriptGenerate")
        self.horizontalLayout_4.addWidget(self.pushButton_GenerateScript_ScriptGenerate, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem1)
        self.verticalLayout_16.addLayout(self.horizontalLayout_4)
        self.gridLayout_11.addLayout(self.verticalLayout_16, 0, 0, 1, 1)
        self.verticalLayout_17 = QtWidgets.QVBoxLayout()
        self.verticalLayout_17.setObjectName("verticalLayout_17")
        self.textBrowser_ScriptGenerate = QtWidgets.QTextBrowser(self.ScriptGenerateTab)
        self.textBrowser_ScriptGenerate.setObjectName("textBrowser_ScriptGenerate")
        self.verticalLayout_17.addWidget(self.textBrowser_ScriptGenerate)
        self.label_18 = QtWidgets.QLabel(self.ScriptGenerateTab)
        self.label_18.setObjectName("label_18")
        self.verticalLayout_17.addWidget(self.label_18, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout_11.addLayout(self.verticalLayout_17, 1, 0, 1, 1)
        self.Maintab.addTab(self.ScriptGenerateTab, "")
        self.TestManagementTab = QtWidgets.QWidget()
        self.TestManagementTab.setObjectName("TestManagementTab")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.TestManagementTab)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.horizontalLayout_9 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_9.setObjectName("horizontalLayout_9")
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_9.addItem(spacerItem2)
        self.label_WorkbookName_TestManagement = QtWidgets.QLabel(self.TestManagementTab)
        self.label_WorkbookName_TestManagement.setStyleSheet("font: 75 10pt \"Arial\";")
        self.label_WorkbookName_TestManagement.setObjectName("label_WorkbookName_TestManagement")
        self.horizontalLayout_9.addWidget(self.label_WorkbookName_TestManagement, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.label_WorkbookName1_TestManagement = QtWidgets.QLabel(self.TestManagementTab)
        self.label_WorkbookName1_TestManagement.setStyleSheet("font: 75 10pt \"Arial\";\n"
"")
        self.label_WorkbookName1_TestManagement.setObjectName("label_WorkbookName1_TestManagement")
        self.horizontalLayout_9.addWidget(self.label_WorkbookName1_TestManagement, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.pushButton_Open_TestManagement = QtWidgets.QPushButton(self.TestManagementTab)
        self.pushButton_Open_TestManagement.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestManagement.setObjectName("pushButton_Open_TestManagement")
        self.horizontalLayout_9.addWidget(self.pushButton_Open_TestManagement, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.pushButton_Open_TestManagement_2 = QtWidgets.QPushButton(self.TestManagementTab)
        self.pushButton_Open_TestManagement_2.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestManagement_2.setObjectName("pushButton_Open_TestManagement_2")
        self.horizontalLayout_9.addWidget(self.pushButton_Open_TestManagement_2, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_9.addItem(spacerItem3)
        self.verticalLayout_5.addLayout(self.horizontalLayout_9)
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.checkBoxCreateDefectsAutomatically = QtWidgets.QCheckBox(self.TestManagementTab)
        self.checkBoxCreateDefectsAutomatically.setStyleSheet("font: 75 10pt \"Arial\";")
        self.checkBoxCreateDefectsAutomatically.setObjectName("checkBoxCreateDefectsAutomatically")
        self.verticalLayout_3.addWidget(self.checkBoxCreateDefectsAutomatically, 0, QtCore.Qt.AlignHCenter)
        self.label_6 = QtWidgets.QLabel(self.TestManagementTab)
        self.label_6.setText("")
        self.label_6.setObjectName("label_6")
        self.verticalLayout_3.addWidget(self.label_6)
        self.verticalLayout_5.addLayout(self.verticalLayout_3)
        self.horizontalLayout_10 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_10.setObjectName("horizontalLayout_10")
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_10.addItem(spacerItem4)
        self.label_ChoosetheTestManagementTool = QtWidgets.QLabel(self.TestManagementTab)
        self.label_ChoosetheTestManagementTool.setObjectName("label_ChoosetheTestManagementTool")
        self.horizontalLayout_10.addWidget(self.label_ChoosetheTestManagementTool, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.radioButton_ALM = QtWidgets.QRadioButton(self.TestManagementTab)
        self.radioButton_ALM.setStyleSheet("font: 75 10pt \"Arial\";")
        self.radioButton_ALM.setObjectName("radioButton_ALM")
        self.horizontalLayout_10.addWidget(self.radioButton_ALM, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.radioButton_JIRA = QtWidgets.QRadioButton(self.TestManagementTab)
        self.radioButton_JIRA.setStyleSheet("font: 75 10pt \"Arial\";")
        self.radioButton_JIRA.setObjectName("radioButton_JIRA")
        self.horizontalLayout_10.addWidget(self.radioButton_JIRA, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.radioButton_NA = QtWidgets.QRadioButton(self.TestManagementTab)
        self.radioButton_NA.setStyleSheet("font: 75 10pt \"Arial\";")
        self.radioButton_NA.setChecked(True)
        self.radioButton_NA.setObjectName("radioButton_NA")
        self.horizontalLayout_10.addWidget(self.radioButton_NA, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_10.addItem(spacerItem5)
        self.verticalLayout_5.addLayout(self.horizontalLayout_10)
        self.verticalLayout_26 = QtWidgets.QVBoxLayout()
        self.verticalLayout_26.setObjectName("verticalLayout_26")
        self.textBrowser_9 = QtWidgets.QTextBrowser(self.TestManagementTab)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(9)
        self.textBrowser_9.setFont(font)
        self.textBrowser_9.setObjectName("textBrowser_9")
        self.verticalLayout_26.addWidget(self.textBrowser_9)
        self.label_19 = QtWidgets.QLabel(self.TestManagementTab)
        self.label_19.setObjectName("label_19")
        self.verticalLayout_26.addWidget(self.label_19, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout_5.addLayout(self.verticalLayout_26)
        self.Maintab.addTab(self.TestManagementTab, "")
        self.TestSetsTab = QtWidgets.QWidget()
        self.TestSetsTab.setObjectName("TestSetsTab")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.TestSetsTab)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.horizontalLayout_15 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_15.setObjectName("horizontalLayout_15")
        spacerItem6 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_15.addItem(spacerItem6)
        self.label_MasterWorkBookName = QtWidgets.QLabel(self.TestSetsTab)
        self.label_MasterWorkBookName.setObjectName("label_MasterWorkBookName")
        self.horizontalLayout_15.addWidget(self.label_MasterWorkBookName, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.label_MasterWorkBookName_2 = QtWidgets.QLabel(self.TestSetsTab)
        self.label_MasterWorkBookName_2.setObjectName("label_MasterWorkBookName_2")
        self.horizontalLayout_15.addWidget(self.label_MasterWorkBookName_2, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.TestSetsOpenBtn_2 = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsOpenBtn_2.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsOpenBtn_2.setObjectName("TestSetsOpenBtn_2")
        self.horizontalLayout_15.addWidget(self.TestSetsOpenBtn_2, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.TestSetsSaveBtn_2 = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsSaveBtn_2.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsSaveBtn_2.setObjectName("TestSetsSaveBtn_2")
        self.horizontalLayout_15.addWidget(self.TestSetsSaveBtn_2, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.TestSetsSaveAsBtn = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsSaveAsBtn.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsSaveAsBtn.setObjectName("TestSetsSaveAsBtn")
        self.horizontalLayout_15.addWidget(self.TestSetsSaveAsBtn, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.TestSetsAddNewBtn = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsAddNewBtn.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsAddNewBtn.setObjectName("TestSetsAddNewBtn")
        self.horizontalLayout_15.addWidget(self.TestSetsAddNewBtn, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.TestSetsAddRowBtn = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsAddRowBtn.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsAddRowBtn.setObjectName("TestSetsAddRowBtn")
        self.horizontalLayout_15.addWidget(self.TestSetsAddRowBtn, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.TestSetsDeleteRowBtn = QtWidgets.QPushButton(self.TestSetsTab)
        self.TestSetsDeleteRowBtn.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.TestSetsDeleteRowBtn.setObjectName("TestSetsDeleteRowBtn")
        self.horizontalLayout_15.addWidget(self.TestSetsDeleteRowBtn, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        spacerItem7 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_15.addItem(spacerItem7)
        self.gridLayout_5.addLayout(self.horizontalLayout_15, 0, 0, 1, 1)
        self.verticalLayout_25 = QtWidgets.QVBoxLayout()
        self.verticalLayout_25.setObjectName("verticalLayout_25")
        self.textBrowser_TestSets = QtWidgets.QTextBrowser(self.TestSetsTab)
        self.textBrowser_TestSets.setObjectName("textBrowser_TestSets")
        self.verticalLayout_25.addWidget(self.textBrowser_TestSets)
        self.gridLayout_5.addLayout(self.verticalLayout_25, 1, 0, 1, 1)
        self.label_20 = QtWidgets.QLabel(self.TestSetsTab)
        self.label_20.setObjectName("label_20")
        self.gridLayout_5.addWidget(self.label_20, 2, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.Maintab.addTab(self.TestSetsTab, "")
        self.TestScriptsTab = QtWidgets.QWidget()
        self.TestScriptsTab.setObjectName("TestScriptsTab")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.TestScriptsTab)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout()
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout()
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.horizontalLayout_20 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_20.setObjectName("horizontalLayout_20")
        spacerItem8 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_20.addItem(spacerItem8)
        self.label_SelecttheScriptSheet = QtWidgets.QLabel(self.TestScriptsTab)
        self.label_SelecttheScriptSheet.setObjectName("label_SelecttheScriptSheet")
        self.horizontalLayout_20.addWidget(self.label_SelecttheScriptSheet, 0, QtCore.Qt.AlignHCenter)
        self.comboBox_TestScripts = QtWidgets.QComboBox(self.TestScriptsTab)
        self.comboBox_TestScripts.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestScripts.setObjectName("comboBox_TestScripts")
        self.horizontalLayout_20.addWidget(self.comboBox_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_Open_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_Open_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestScripts.setObjectName("pushButton_Open_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_Open_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_Save_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_Save_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_TestScripts.setObjectName("pushButton_Save_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_Save_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_SaveAs_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_SaveAs_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_SaveAs_TestScripts.setObjectName("pushButton_SaveAs_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_SaveAs_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_AddNew_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_AddNew_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddNew_TestScripts.setObjectName("pushButton_AddNew_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_AddNew_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_AddRow_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_AddRow_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddRow_TestScripts.setObjectName("pushButton_AddRow_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_AddRow_TestScripts, 0, QtCore.Qt.AlignHCenter)
        self.pushButton_deleteRow_TestScripts = QtWidgets.QPushButton(self.TestScriptsTab)
        self.pushButton_deleteRow_TestScripts.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_deleteRow_TestScripts.setObjectName("pushButton_deleteRow_TestScripts")
        self.horizontalLayout_20.addWidget(self.pushButton_deleteRow_TestScripts, 0, QtCore.Qt.AlignHCenter)
        spacerItem9 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_20.addItem(spacerItem9)
        self.verticalLayout_6.addLayout(self.horizontalLayout_20)
        self.verticalLayout_4.addLayout(self.verticalLayout_6)
        self.verticalLayout_7 = QtWidgets.QVBoxLayout()
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.horizontalLayout_21 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_21.setObjectName("horizontalLayout_21")
        spacerItem10 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_21.addItem(spacerItem10)
        self.label_TestAutomationToolAndTechnology = QtWidgets.QLabel(self.TestScriptsTab)
        self.label_TestAutomationToolAndTechnology.setObjectName("label_TestAutomationToolAndTechnology")
        self.horizontalLayout_21.addWidget(self.label_TestAutomationToolAndTechnology)
        self.comboBox_tool = QtWidgets.QComboBox(self.TestScriptsTab)
        self.comboBox_tool.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_tool.setObjectName("comboBox_tool")
        self.comboBox_tool.addItem("")
        self.horizontalLayout_21.addWidget(self.comboBox_tool)
        self.comboBox_technology = QtWidgets.QComboBox(self.TestScriptsTab)
        self.comboBox_technology.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_technology.setObjectName("comboBox_technology")
        self.comboBox_technology.addItem("")
        self.comboBox_technology.addItem("")
        self.comboBox_technology.setItemText(1, "")
        self.horizontalLayout_21.addWidget(self.comboBox_technology)
        spacerItem11 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_21.addItem(spacerItem11)
        self.verticalLayout_7.addLayout(self.horizontalLayout_21)
        self.verticalLayout_4.addLayout(self.verticalLayout_7)
        self.gridLayout_3.addLayout(self.verticalLayout_4, 0, 0, 1, 1)
        self.verticalLayout_24 = QtWidgets.QVBoxLayout()
        self.verticalLayout_24.setObjectName("verticalLayout_24")
        self.textBrowser_7 = QtWidgets.QTextBrowser(self.TestScriptsTab)
        self.textBrowser_7.setObjectName("textBrowser_7")
        self.verticalLayout_24.addWidget(self.textBrowser_7)
        self.gridLayout_3.addLayout(self.verticalLayout_24, 1, 0, 1, 1)
        self.label_21 = QtWidgets.QLabel(self.TestScriptsTab)
        self.label_21.setObjectName("label_21")
        self.gridLayout_3.addWidget(self.label_21, 2, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.Maintab.addTab(self.TestScriptsTab, "")
        self.TestComponentsTab = QtWidgets.QWidget()
        self.TestComponentsTab.setObjectName("TestComponentsTab")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.TestComponentsTab)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.horizontalLayout_22 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_22.setObjectName("horizontalLayout_22")
        spacerItem12 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_22.addItem(spacerItem12)
        self.label_SleecttheComponentSheet = QtWidgets.QLabel(self.TestComponentsTab)
        self.label_SleecttheComponentSheet.setObjectName("label_SleecttheComponentSheet")
        self.horizontalLayout_22.addWidget(self.label_SleecttheComponentSheet, 0, QtCore.Qt.AlignVCenter)
        self.comboBox_TestComponents = QtWidgets.QComboBox(self.TestComponentsTab)
        self.comboBox_TestComponents.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestComponents.setObjectName("comboBox_TestComponents")
        self.horizontalLayout_22.addWidget(self.comboBox_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_Open_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_Open_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestComponents.setObjectName("pushButton_Open_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_Open_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_Save_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_Save_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_TestComponents.setObjectName("pushButton_Save_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_Save_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_SaveAs_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_SaveAs_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_SaveAs_TestComponents.setObjectName("pushButton_SaveAs_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_SaveAs_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddNew_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_AddNew_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddNew_TestComponents.setObjectName("pushButton_AddNew_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_AddNew_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddRow_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_AddRow_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddRow_TestComponents.setObjectName("pushButton_AddRow_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_AddRow_TestComponents, 0, QtCore.Qt.AlignTop)
        self.pushButton_DeleteRow_TestComponents = QtWidgets.QPushButton(self.TestComponentsTab)
        self.pushButton_DeleteRow_TestComponents.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_DeleteRow_TestComponents.setObjectName("pushButton_DeleteRow_TestComponents")
        self.horizontalLayout_22.addWidget(self.pushButton_DeleteRow_TestComponents, 0, QtCore.Qt.AlignTop)
        spacerItem13 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_22.addItem(spacerItem13)
        self.gridLayout_6.addLayout(self.horizontalLayout_22, 0, 0, 1, 1)
        self.verticalLayout_22 = QtWidgets.QVBoxLayout()
        self.verticalLayout_22.setObjectName("verticalLayout_22")
        spacerItem13 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.textBrowser_TestComponents = QtWidgets.QTextBrowser(self.TestComponentsTab)
        self.textBrowser_TestComponents.setObjectName("textBrowser_TestComponents")
        self.verticalLayout_22.addWidget(self.textBrowser_TestComponents)
        self.gridLayout_6.addLayout(self.verticalLayout_22, 1, 0, 1, 1)
        self.label_22 = QtWidgets.QLabel(self.TestComponentsTab)
        self.label_22.setObjectName("label_22")
        self.gridLayout_6.addWidget(self.label_22, 2, 0, 1, 1, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.Maintab.addTab(self.TestComponentsTab, "")
        self.TestDataTab = QtWidgets.QWidget()
        self.TestDataTab.setObjectName("TestDataTab")
        self.gridLayout_7 = QtWidgets.QGridLayout(self.TestDataTab)
        self.gridLayout_7.setObjectName("gridLayout_7")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        spacerItem14 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem14)
        self.label_SelecttheTestDataSheet = QtWidgets.QLabel(self.TestDataTab)
        self.label_SelecttheTestDataSheet.setObjectName("label_SelecttheTestDataSheet")
        self.horizontalLayout.addWidget(self.label_SelecttheTestDataSheet, 0, QtCore.Qt.AlignVCenter)
        self.comboBox_TestData = QtWidgets.QComboBox(self.TestDataTab)
        self.comboBox_TestData.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestData.setObjectName("comboBox_TestData")
        self.horizontalLayout.addWidget(self.comboBox_TestData, 0, QtCore.Qt.AlignVCenter)
        self.pushButton_Open_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_Open_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestData.setObjectName("pushButton_Open_TestData")
        self.horizontalLayout.addWidget(self.pushButton_Open_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_Save_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_Save_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_TestData.setObjectName("pushButton_Save_TestData")
        self.horizontalLayout.addWidget(self.pushButton_Save_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_SaveAs_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_SaveAs_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_SaveAs_TestData.setObjectName("pushButton_SaveAs_TestData")
        self.horizontalLayout.addWidget(self.pushButton_SaveAs_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddNew_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_AddNew_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddNew_TestData.setObjectName("pushButton_AddNew_TestData")
        self.horizontalLayout.addWidget(self.pushButton_AddNew_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddRow_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_AddRow_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddRow_TestData.setObjectName("pushButton_AddRow_TestData")
        self.horizontalLayout.addWidget(self.pushButton_AddRow_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_DeleteRow_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_DeleteRow_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_DeleteRow_TestData.setObjectName("pushButton_DeleteRow_TestData")
        self.horizontalLayout.addWidget(self.pushButton_DeleteRow_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddColumn_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_AddColumn_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddColumn_TestData.setObjectName("pushButton_AddColumn_TestData")
        self.horizontalLayout.addWidget(self.pushButton_AddColumn_TestData, 0, QtCore.Qt.AlignTop)
        self.pushButton_DeleteColumn_TestData = QtWidgets.QPushButton(self.TestDataTab)
        self.pushButton_DeleteColumn_TestData.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_DeleteColumn_TestData.setObjectName("pushButton_DeleteColumn_TestData")
        self.horizontalLayout.addWidget(self.pushButton_DeleteColumn_TestData, 0, QtCore.Qt.AlignTop)
        spacerItem15 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem15)
        self.gridLayout_7.addLayout(self.horizontalLayout, 0, 0, 1, 1)
        self.verticalLayout_21 = QtWidgets.QVBoxLayout()
        self.verticalLayout_21.setObjectName("verticalLayout_21")
        self.textBrowser_TestData = QtWidgets.QTextBrowser(self.TestDataTab)
        self.textBrowser_TestData.setObjectName("textBrowser_TestData")
        self.verticalLayout_21.addWidget(self.textBrowser_TestData)
        self.gridLayout_7.addLayout(self.verticalLayout_21, 1, 0, 1, 1)
        self.label_23 = QtWidgets.QLabel(self.TestDataTab)
        self.label_23.setObjectName("label_23")
        self.gridLayout_7.addWidget(self.label_23, 2, 0, 1, 1, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.Maintab.addTab(self.TestDataTab, "")
        self.TestObjectsTab = QtWidgets.QWidget()
        self.TestObjectsTab.setObjectName("TestObjectsTab")
        self.gridLayout_8 = QtWidgets.QGridLayout(self.TestObjectsTab)
        self.gridLayout_8.setObjectName("gridLayout_8")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        spacerItem16 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem16)
        self.label_SelecttheObjectsExcel = QtWidgets.QLabel(self.TestObjectsTab)
        self.label_SelecttheObjectsExcel.setObjectName("label_SelecttheObjectsExcel")
        self.horizontalLayout_2.addWidget(self.label_SelecttheObjectsExcel, 0, QtCore.Qt.AlignVCenter)
        self.comboBox_TestObjects = QtWidgets.QComboBox(self.TestObjectsTab)
        self.comboBox_TestObjects.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestObjects.setObjectName("comboBox_TestObjects")
        self.horizontalLayout_2.addWidget(self.comboBox_TestObjects, 0, QtCore.Qt.AlignVCenter)
        self.pushButton_Open_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_Open_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Open_TestObjects.setObjectName("pushButton_Open_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_Open_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_Save_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_Save_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_TestObjects.setObjectName("pushButton_Save_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_Save_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_SaveAs_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_SaveAs_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_SaveAs_TestObjects.setObjectName("pushButton_SaveAs_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_SaveAs_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddNew_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_AddNew_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddNew_TestObjects.setObjectName("pushButton_AddNew_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_AddNew_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddRow_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_AddRow_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddRow_TestObjects.setObjectName("pushButton_AddRow_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_AddRow_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_DeleteRow_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_DeleteRow_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_DeleteRow_TestObjects.setObjectName("pushButton_DeleteRow_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_DeleteRow_TestObjects, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddColumn_TestObjects = QtWidgets.QPushButton(self.TestObjectsTab)
        self.pushButton_AddColumn_TestObjects.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddColumn_TestObjects.setObjectName("pushButton_AddColumn_TestObjects")
        self.horizontalLayout_2.addWidget(self.pushButton_AddColumn_TestObjects, 0, QtCore.Qt.AlignTop)
        spacerItem17 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem17)
        self.gridLayout_8.addLayout(self.horizontalLayout_2, 0, 0, 1, 1)
        self.verticalLayout_20 = QtWidgets.QVBoxLayout()
        self.verticalLayout_20.setObjectName("verticalLayout_20")
        self.textBrowser_TestObjects = QtWidgets.QTextBrowser(self.TestObjectsTab)
        self.textBrowser_TestObjects.setObjectName("textBrowser_TestObjects")
        self.verticalLayout_20.addWidget(self.textBrowser_TestObjects)
        self.gridLayout_8.addLayout(self.verticalLayout_20, 1, 0, 1, 1)
        self.label_24 = QtWidgets.QLabel(self.TestObjectsTab)
        self.label_24.setObjectName("label_24")
        self.gridLayout_8.addWidget(self.label_24, 2, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.Maintab.addTab(self.TestObjectsTab, "")
        self.TestExecutionTab = QtWidgets.QWidget()
        self.TestExecutionTab.setEnabled(True)
        self.TestExecutionTab.setObjectName("TestExecutionTab")
        self.gridLayout = QtWidgets.QGridLayout(self.TestExecutionTab)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout_30 = QtWidgets.QVBoxLayout()
        self.verticalLayout_30.setObjectName("verticalLayout_30")
        self.textBrowser_TestExecution = QtWidgets.QTextBrowser(self.TestExecutionTab)
        self.textBrowser_TestExecution.setObjectName("textBrowser_TestExecution")
        self.verticalLayout_30.addWidget(self.textBrowser_TestExecution)
        # self.verticalLayout_31 = QtWidgets.QVBoxLayout()
        # self.verticalLayout_31.setObjectName("verticalLayout_31")
        # self.verticalLayout_30.addLayout(self.verticalLayout_31)
        self.gridLayout.addLayout(self.verticalLayout_30, 1, 0, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.TestExecutionTab)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 2, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.verticalLayout_27 = QtWidgets.QVBoxLayout()
        self.verticalLayout_27.setObjectName("verticalLayout_27")
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        spacerItem18 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_5.addItem(spacerItem18)
        self.pushButton_OpenDriver_TestExecution = QtWidgets.QPushButton(self.TestExecutionTab)
        self.pushButton_OpenDriver_TestExecution.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_OpenDriver_TestExecution.setObjectName("pushButton_OpenDriver_TestExecution")
        self.horizontalLayout_5.addWidget(self.pushButton_OpenDriver_TestExecution, 0, QtCore.Qt.AlignTop)
        self.pushButton_Save_TestExecution = QtWidgets.QPushButton(self.TestExecutionTab)
        self.pushButton_Save_TestExecution.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Save_TestExecution.setObjectName("pushButton_Save_TestExecution")
        self.horizontalLayout_5.addWidget(self.pushButton_Save_TestExecution, 0, QtCore.Qt.AlignTop)
        self.label_TestExecutionTool = QtWidgets.QLabel(self.TestExecutionTab)
        self.label_TestExecutionTool.setStyleSheet("\n"
"\n"
"")
        self.label_TestExecutionTool.setObjectName("label_TestExecutionTool")
        self.horizontalLayout_5.addWidget(self.label_TestExecutionTool, 0, QtCore.Qt.AlignHCenter|QtCore.Qt.AlignVCenter)
        self.comboBox_TestExecution = QtWidgets.QComboBox(self.TestExecutionTab)
        self.comboBox_TestExecution.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestExecution.setObjectName("comboBox_TestExecution")
        self.comboBox_TestExecution.addItems(self.parser.get('settings','tools').split(','))
        self.horizontalLayout_5.addWidget(self.comboBox_TestExecution, 0, QtCore.Qt.AlignVCenter)
        self.pushButton_Execute_TestExecution = QtWidgets.QPushButton(self.TestExecutionTab)
        self.pushButton_Execute_TestExecution.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_Execute_TestExecution.setObjectName("pushButton_Execute_TestExecution")
        self.horizontalLayout_5.addWidget(self.pushButton_Execute_TestExecution, 0, QtCore.Qt.AlignTop)
        self.pushButton_AddRow_TestExecution = QtWidgets.QPushButton(self.TestExecutionTab)
        self.pushButton_AddRow_TestExecution.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_AddRow_TestExecution.setObjectName("pushButton_AddRow_TestExecution")
        self.horizontalLayout_5.addWidget(self.pushButton_AddRow_TestExecution, 0, QtCore.Qt.AlignTop)
        self.pushButton_DeleteRow_TestExecution = QtWidgets.QPushButton(self.TestExecutionTab)
        self.pushButton_DeleteRow_TestExecution.setStyleSheet("QPushButton {\n"
"background-color: qradialgradient(spread:pad, cx:0.5, cy:0.5, radius:0.5, fx:0.511364, fy:0.534136, stop:0.568182 rgba(255, 255, 255, 211), stop:1 rgba(148, 209, 255, 137));\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 6px;\n"
"border-color: black;\n"
"min-width: 6em;\n"
"padding: 6px;\n"
"}\n"
"\n"
"\n"
"QPushButton:pressed {\n"
"    background-color: rgba(255, 255, 255, 211);\n"
"    \n"
"}\n"
"")
        self.pushButton_DeleteRow_TestExecution.setObjectName("pushButton_DeleteRow_TestExecution")
        self.horizontalLayout_5.addWidget(self.pushButton_DeleteRow_TestExecution, 0, QtCore.Qt.AlignTop)
        spacerItem19 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_5.addItem(spacerItem19)
        self.verticalLayout_27.addLayout(self.horizontalLayout_5)
        self.gridLayout.addLayout(self.verticalLayout_27, 0, 0, 1, 1)
        self.Maintab.addTab(self.TestExecutionTab, "")

        #*************************************************************************************
        self.TestReportsTab = QtWidgets.QWidget()
        self.TestReportsTab.setObjectName("TestReportsTab")
        self.gridLayout_9 = QtWidgets.QGridLayout(self.TestReportsTab)
        self.gridLayout_9.setObjectName("gridLayout_9")
        self.verticalLayout_13 = QtWidgets.QVBoxLayout()
        self.verticalLayout_13.setObjectName("verticalLayout_13")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_TestResultPath = QtWidgets.QLabel(self.TestReportsTab)
        self.label_TestResultPath.setObjectName("label_TestResultPath")
        self.horizontalLayout_3.addWidget(self.label_TestResultPath, 0, QtCore.Qt.AlignRight|QtCore.Qt.AlignVCenter)
        self.comboBox_TestReports = QtWidgets.QComboBox(self.TestReportsTab)
        self.comboBox_TestReports.setStyleSheet("  border: 1px solid gray;\n"
"    border-radius: 3px;\n"
"    padding: 1px 18px 1px 3px;\n"
"    min-width: 5em;")
        self.comboBox_TestReports.setObjectName("comboBox_TestReports")
        self.horizontalLayout_3.addWidget(self.comboBox_TestReports, 0, QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.verticalLayout_13.addLayout(self.horizontalLayout_3)
        self.gridLayout_9.addLayout(self.verticalLayout_13, 0, 0, 1, 1)
        self.verticalLayout_14 = QtWidgets.QVBoxLayout()
        self.verticalLayout_14.setObjectName("verticalLayout_14")
        self.textBrowser_ExecutionSummary = QtWidgets.QTextBrowser(self.TestReportsTab)
        self.textBrowser_ExecutionSummary.setObjectName("textBrowser_ExecutionSummary")
        self.verticalLayout_14.addWidget(self.textBrowser_ExecutionSummary)
        self.gridLayout_9.addLayout(self.verticalLayout_14, 2, 0, 1, 1)
        self.verticalLayout_15 = QtWidgets.QVBoxLayout()
        self.verticalLayout_15.setObjectName("verticalLayout_15")
        self.textBrowser_defectSummary = QtWidgets.QTextBrowser(self.TestReportsTab)
        self.textBrowser_defectSummary.setObjectName("textBrowser_defectSummary")
        self.verticalLayout_15.addWidget(self.textBrowser_defectSummary)
        self.label_15 = QtWidgets.QLabel(self.TestReportsTab)
        self.label_15.setText("")
        self.label_15.setObjectName("label_15")
        self.verticalLayout_15.addWidget(self.label_15)
        self.gridLayout_9.addLayout(self.verticalLayout_15, 4, 0, 1, 1)
        self.label_defectSummary = QtWidgets.QLabel(self.TestReportsTab)
        self.label_defectSummary.setObjectName("label_defectSummary")
        self.gridLayout_9.addWidget(self.label_defectSummary, 3, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_ExecutionSummary = QtWidgets.QLabel(self.TestReportsTab)
        self.label_ExecutionSummary.setObjectName("label_ExecutionSummary")
        self.gridLayout_9.addWidget(self.label_ExecutionSummary, 1, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_25 = QtWidgets.QLabel(self.TestReportsTab)
        self.label_25.setObjectName("label_25")
        self.gridLayout_9.addWidget(self.label_25, 5, 0, 1, 1, QtCore.Qt.AlignHCenter)
        self.Maintab.addTab(self.TestReportsTab, "")
        self.AboutTab = QtWidgets.QWidget()
        self.AboutTab.setObjectName("AboutTab")
        self.gridLayout_12 = QtWidgets.QGridLayout(self.AboutTab)
        self.gridLayout_12.setObjectName("gridLayout_12")
        self.label_10 = QtWidgets.QLabel(self.AboutTab)
        self.label_10.setObjectName("label_10")
        self.gridLayout_12.addWidget(self.label_10, 0, 0, 1, 1)
        self.Maintab.addTab(self.AboutTab, "")
        self.gridLayout_2.addWidget(self.Maintab, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1900, 18))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.menuFile.setFont(font)
        self.menuFile.setObjectName("menuFile")
        self.menuConfigure = QtWidgets.QMenu(self.menubar)
        self.menuConfigure.setObjectName("menuConfigure")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionConfigure = QtWidgets.QAction(MainWindow)
        self.actionConfigure.setObjectName("actionConfigure")
        self.actionExit = QtWidgets.QAction(MainWindow)
        self.actionExit.setObjectName("actionExit")
        self.menuFile.addAction(self.actionConfigure)
        self.menuFile.addAction(self.actionExit)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuConfigure.menuAction())

        #Test Sets
        self.TestSetsOpenBtn_2.clicked.connect(self.on_click_open_test_sets )
        self.TestSetsSaveBtn_2.clicked.connect(self.on_click_save_test_sets)
        self.TestSetsSaveAsBtn.clicked.connect(self.on_click_save_as_test_sets)

        #Test Scripts
        self.pushButton_Open_TestScripts.clicked.connect(self.on_click_open_test_Scripts)
        self.pushButton_Save_TestScripts.clicked.connect(self.on_click_save_test_scripts)
        self.pushButton_SaveAs_TestScripts.clicked.connect(self.on_click_save_as_test_scripts)
        self.comboBox_TestScripts.activated.connect(self.select_from_comboBox_TestScripts)

        #Test Components
        self.pushButton_Open_TestComponents.clicked.connect(self.on_click_open_test_Components)
        self.pushButton_Save_TestComponents.clicked.connect(self.on_click_save_test_Components)
        self.pushButton_SaveAs_TestComponents.clicked.connect(self.on_click_save_as_test_Components)
        self.comboBox_TestComponents.activated.connect(self.select_from_comboBox_TestComponents)

        #Test Data
        self.pushButton_Open_TestData.clicked.connect(self.on_click_open_test_Data)
        self.pushButton_Save_TestData.clicked.connect(self.on_click_save_test_Data)
        self.pushButton_SaveAs_TestData.clicked.connect(self.on_click_save_as_test_Data)
        self.comboBox_TestData.activated.connect(self.select_from_comboBox_TestData)

        #Test Object
        self.pushButton_Open_TestObjects.clicked.connect(self.on_click_open_test_Objects)
        self.pushButton_Save_TestObjects.clicked.connect(self.on_click_save_test_Objects)
        self.pushButton_SaveAs_TestObjects.clicked.connect(self.on_click_save_as_test_Objects)
        self.comboBox_TestObjects.activated.connect(self.select_from_comboBox_TestObjects)

        #Test Execution
        self.pushButton_OpenDriver_TestExecution.clicked.connect(self.on_click_open_driver_test_execution)
        self.pushButton_Save_TestExecution.clicked.connect(self.on_click_save_test_execution)
        self.pushButton_Execute_TestExecution.clicked.connect(self.on_click_execute_test_execution)


        self.retranslateUi(MainWindow)
        self.Maintab.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def on_click_open_test_sets(self):
            Tk().withdraw()
            self.masterSelectedFile = askopenfilename(initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                                       filetypes=[("Excel files", "Master_*.xlsx .xls")])
            if self.masterSelectedFile :
                item=self.verticalLayout_25.takeAt(0)
                w=item.widget()
                w.deleteLater()
                _translate = QtCore.QCoreApplication.translate
                masterSelectedFileName=basename(self.masterSelectedFile)
                self.label_MasterWorkBookName_2.setText(_translate("MainWindow",masterSelectedFileName))
                self.set_view = QTableView()
                a = self.masterSelectedFile
                df = pd.read_excel(a,sheet_name='Master')
                self.your_pandas_data_sets = df.replace(np.nan, '', regex=True)
                header = list(self.your_pandas_data_sets.columns.values)
                self.set_model = pandatest.PandasModel(self.your_pandas_data_sets, header)
                self.set_view.setModel(self.set_model)
                self.verticalLayout_25.addWidget(self.set_view)
            else:
                    pass

    def on_click_save_test_sets(self):
            if self.masterSelectedFile is  None:
                    Tk().withdraw()
                    messagebox.showinfo("Error", "Please Select the Master Sheet")
            else:
                for row in range(self.set_model.rowCount()):
                    for column in range(self.set_model.columnCount()):
                        index = self.set_model.index(row, column)
                        da = self.set_model.itemData(index)
                        self.your_pandas_data_sets.iat[row, column] = da[0]
                self.your_pandas_data_sets.to_excel(self.masterSelectedFile,index=False)

    def on_click_save_as_test_sets(self):
            if self.masterSelectedFile is  None:
                    Tk().withdraw()
                    messagebox.showinfo("Error", "Please Select the Master Sheet")
            elif self.masterSelectedFile is  not None:
                    Tk().withdraw()
                    FileName = asksaveasfilename(
                            initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                            filetypes=[("Excel files", "Master_*.xlsx .xls")],
                                defaultextension=".xlsx",
                                initialfile="Master_*")
                    if FileName is not None:
                        for row in range(self.set_model.rowCount()):
                            for column in range(self.set_model.columnCount()):
                                index = self.set_model.index(row, column)
                                da = self.set_model.itemData(index)
                                self.your_pandas_data_sets.iat[row, column] = da[0]
                        self.your_pandas_data_sets.to_excel(FileName,index=False)
                    else:
                         pass
            else:
                    pass

    def select_from_comboBox_TestScripts(self):
        selectedSheet = self.comboBox_TestScripts.currentText()
        if self.testScriptSelectedFile:
            wb = load_workbook(self.testScriptSelectedFile)
            item = self.verticalLayout_24.takeAt(0)
            w = item.widget()
            w.deleteLater()
            self.script_sheet = selectedSheet
            self.script_view = QTableView()
            a = self.testScriptSelectedFile
            df = pd.read_excel(a, sheet_name=self.script_sheet)
            self.your_pandas_data_scripts = df.replace(np.nan, '', regex=True)
            header = list(self.your_pandas_data_scripts.columns.values)
            self.script_model = pandatest.PandasModel(self.your_pandas_data_scripts, header)
            self.script_view.setModel(self.script_model)
            self.verticalLayout_24.addWidget(self.script_view)
        else:
            pass

    def on_click_open_test_Scripts(self):
        Tk().withdraw()
        self.testScriptSelectedFile = askopenfilename(
            initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
            filetypes=[("Excel files", "Script_*.xlsx .xls")])
        if self.testScriptSelectedFile:
            wb = load_workbook(self.testScriptSelectedFile)
            sheetNames = wb.sheetnames
            self.comboBox_TestScripts.clear()
            self.comboBox_TestScripts.addItems(sheetNames)
            Dialog = QtWidgets.QDialog()
            ui = Ui_Dialog_ComboBox()
            ui.setupUi(Dialog, sheetNames)
            item = self.verticalLayout_24.takeAt(0)
            w = item.widget()
            w.deleteLater()
            self.script_sheet = ui.comboBox.currentText()
            self.comboBox_TestScripts.setCurrentText(self.script_sheet)
            self.script_view=QTableView()
            a = self.testScriptSelectedFile
            df=pd.read_excel(a,sheet_name=self.script_sheet)
            self.your_pandas_data_scripts = df.replace(np.nan, '', regex=True)
            header=list(self.your_pandas_data_scripts.columns.values)
            self.script_model=pandatest.PandasModel(self.your_pandas_data_scripts,header)
            self.script_view.setModel(self.script_model)
            self.verticalLayout_24.addWidget(self.script_view)
        else:
                    pass

    def on_click_save_test_scripts(self):
        if self.testScriptSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Script Sheet")
        else:
            for row in range(self.script_model.rowCount()):
                for column in range(self.script_model.columnCount()):
                    index = self.script_model.index(row, column)
                    da = self.script_model.itemData(index)
                    self.your_pandas_data_scripts.iat[row,column]=da[0]
            book=load_workbook(self.testScriptSelectedFile)
            writer = ExcelWriter(self.testScriptSelectedFile,engine='openpyxl')
            writer.book=book
            writer.sheets=dict((ws.title, ws) for ws in book.worksheets)
            self.your_pandas_data_scripts.to_excel(writer,self.script_sheet,index=False)
            writer.save()

    def on_click_save_as_test_scripts(self):
        if self.testScriptSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Script Sheet")
        elif self.testScriptSelectedFile is not None:
            Tk().withdraw()
            FileName = asksaveasfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "Script_*.xlsx .xls")],
                defaultextension=".xlsx",
                initialfile="Script_*")
            if FileName is not None:
                for row in range(self.script_model.rowCount()):
                    for column in range(self.script_model.columnCount()):
                        index = self.script_model.index(row, column)
                        da = self.script_model.itemData(index)
                        self.your_pandas_data_scripts.iat[row, column] = da[0]
                writer = ExcelWriter(FileName)
                self.your_pandas_data_scripts.to_excel(writer, self.script_sheet,index=False)
                writer.save()
            else:
                pass
        else:
            pass

    def select_from_comboBox_TestComponents(self):
        selectedSheet=self.comboBox_TestComponents.currentText()
        if self.testComponentSelectedFile:
            wb = load_workbook(self.testComponentSelectedFile)
            item = self.verticalLayout_22.takeAt(0)
            w = item.widget()
            w.deleteLater()
            self.Component_Sheet = selectedSheet
            self.component_view = QTableView()
            a = self.testComponentSelectedFile
            df = pd.read_excel(a, sheet_name=self.Component_Sheet)
            self.your_pandas_data_components = df.replace(np.nan, '', regex=True)
            header = list(self.your_pandas_data_components.columns.values)
            self.component_model = pandatest.PandasModel(self.your_pandas_data_components, header)
            self.component_view.setModel(self.component_model)
            self.verticalLayout_22.addWidget(self.component_view)
        else:
            pass

    def on_click_open_test_Components(self):
            Tk().withdraw()
            self.testComponentSelectedFile = askopenfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "Reusable_Components*.xlsx .xls")])
            if self.testComponentSelectedFile:
                wb = load_workbook(self.testComponentSelectedFile)
                sheetNames = wb.sheetnames
                print(sheetNames)
                self.comboBox_TestComponents.clear()
                self.comboBox_TestComponents.addItems(sheetNames)
                Dialog = QtWidgets.QDialog()
                ui = Ui_Dialog_ComboBox()
                ui.setupUi(Dialog, sheetNames)
                item = self.verticalLayout_22.takeAt(0)
                w = item.widget()
                w.deleteLater()
                self.Component_Sheet = ui.comboBox.currentText()
                self.comboBox_TestComponents.setCurrentText(self.Component_Sheet)
                self.component_view = QTableView()
                a = self.testComponentSelectedFile
                df = pd.read_excel(a, sheet_name=self.Component_Sheet)
                self.your_pandas_data_components = df.replace(np.nan, '', regex=True)
                header = list(self.your_pandas_data_components.columns.values)
                self.component_model = pandatest.PandasModel(self.your_pandas_data_components, header)
                self.component_view.setModel(self.component_model)
                self.verticalLayout_22.addWidget(self.component_view)
            else:
                pass
    def on_click_save_test_Components(self):
        if self.testComponentSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Component Sheet")
        else:
            for row in range(self.component_model.rowCount()):
                for column in range(self.component_model.columnCount()):
                    index = self.component_model.index(row, column)
                    da = self.component_model.itemData(index)
                    self.your_pandas_data_components.iat[row, column] = da[0]
            book = load_workbook(self.testComponentSelectedFile)
            writer = ExcelWriter(self.testComponentSelectedFile, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            self.your_pandas_data_components.to_excel(writer, self.Component_Sheet,index=False)
            writer.save()

    def on_click_save_as_test_Components(self):
        if self.testComponentSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Component Sheet")
        elif self.testComponentSelectedFile is not None:
            Tk().withdraw()
            FileName = asksaveasfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "Reusable_*.xlsx .xls")],
                defaultextension=".xlsx",
                initialfile="Reusable_*")
            if FileName is not None:
                for row in range(self.component_model.rowCount()):
                    for column in range(self.component_model.columnCount()):
                        index = self.component_model.index(row, column)
                        da = self.component_model.itemData(index)
                        self.your_pandas_data_components.iat[row, column] = da[0]
                writer = ExcelWriter(FileName)
                self.your_pandas_data_components.to_excel(writer, self.Component_Sheet,index=False)
                writer.save()
            else:
                pass
        else:
            pass

    def on_click_open_test_Data(self):
            Tk().withdraw()
            self.testDataSelectedFile = askopenfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "TestData_*.xlsx .xls")])
            if self.testDataSelectedFile:
                wb = load_workbook(self.testDataSelectedFile)
                sheetNames = wb.sheetnames
                print(sheetNames)
                self.comboBox_TestData.clear()
                self.comboBox_TestData.addItems(sheetNames)
                Dialog = QtWidgets.QDialog()
                ui = Ui_Dialog_ComboBox()
                ui.setupUi(Dialog, sheetNames)
                item = self.verticalLayout_21.takeAt(0)
                w = item.widget()
                w.deleteLater()
                self.Data_sheet = ui.comboBox.currentText()
                self.comboBox_TestData.setCurrentText(self.Data_sheet)
                self.data_view = QTableView()
                a = self.testDataSelectedFile
                df = pd.read_excel(a, sheet_name=self.Data_sheet)
                self.your_pandas_data_data = df.replace(np.nan, '', regex=True)
                header = list(self.your_pandas_data_data.columns.values)
                self.data_model = pandatest.PandasModel(self.your_pandas_data_data, header)
                self.data_view.setModel(self.data_model)
                self.verticalLayout_21.addWidget(self.data_view)
            else:
                pass

    def on_click_save_test_Data(self):
        if self.testDataSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Data Sheet")
        else:
            for row in range(self.data_model.rowCount()):
                for column in range(self.data_model.columnCount()):
                    index = self.data_model.index(row, column)
                    da = self.data_model.itemData(index)
                    self.your_pandas_data_data.iat[row, column] = da[0]
            book = load_workbook(self.testDataSelectedFile)
            writer = ExcelWriter(self.testDataSelectedFile, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            self.your_pandas_data_data.to_excel(writer, self.Data_sheet,index=False)
            writer.save()

    def on_click_save_as_test_Data(self):
        if self.testDataSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Data Sheet")
        elif self.testDataSelectedFile is not None:
            Tk().withdraw()
            FileName = asksaveasfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "TestData_*.xlsx .xls")],
                defaultextension=".xlsx",
                initialfile="TestData_*")
            if FileName is not None:
                for row in range(self.data_model.rowCount()):
                    for column in range(self.data_model.columnCount()):
                        index = self.data_model.index(row, column)
                        da = self.data_model.itemData(index)
                        self.your_pandas_data_data.iat[row, column] = da[0]
                writer = ExcelWriter(FileName)
                self.your_pandas_data_data.to_excel(writer, self.Data_sheet,index=False)
                writer.save()
            else:
                pass
        else:
            pass

    def select_from_comboBox_TestData(self):
        selectedSheet=self.comboBox_TestData.currentText()
        if self.testDataSelectedFile:
            wb = load_workbook(self.testDataSelectedFile)
            item = self.verticalLayout_21.takeAt(0)
            w = item.widget()
            w.deleteLater()
            self.Data_sheet = selectedSheet
            self.data_view = QTableView()
            a = self.testDataSelectedFile
            df = pd.read_excel(a, sheet_name=self.Data_sheet)
            self.your_pandas_data_data = df.replace(np.nan, '', regex=True)
            header = list(self.your_pandas_data_data.columns.values)
            self.data_model = pandatest.PandasModel(self.your_pandas_data_data, header)
            self.data_view.setModel(self.data_model)
            self.verticalLayout_21.addWidget(self.data_view)
        else:
            pass

    def on_click_open_test_Objects(self):
            Tk().withdraw()
            self.testObjectsSelectedFile = askopenfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "Script_*.xlsx .xls")])
            if self.testObjectsSelectedFile:
                wb = load_workbook(self.testObjectsSelectedFile)
                sheetNames = wb.sheetnames
                print(sheetNames)
                self.comboBox_TestObjects.clear()
                self.comboBox_TestObjects.addItems(sheetNames)
                Dialog = QtWidgets.QDialog()
                ui = Ui_Dialog_ComboBox()
                ui.setupUi(Dialog, sheetNames)
                item = self.verticalLayout_20.takeAt(0)
                print(item)
                w = item.widget()
                print(w)
                w.deleteLater()
                self.Object_sheet = ui.comboBox.currentText()
                self.comboBox_TestObjects.setCurrentText(self.Object_sheet)
                self.object_view = QTableView()
                a = self.testObjectsSelectedFile
                df = pd.read_excel(a, sheet_name=self.Object_sheet)
                self.your_pandas_data_objects = df.replace(np.nan, '', regex=True)
                header = list(self.your_pandas_data_objects.columns.values)
                self.object_model = pandatest.PandasModel(self.your_pandas_data_objects, header)
                self.object_view.setModel(self.object_model)
                self.verticalLayout_20.addWidget(self.object_view)
            else:
                pass

    def on_click_save_test_Objects(self):
        if self.testObjectsSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Object Sheet")
        else:
            for row in range(self.object_model.rowCount()):
                for column in range(self.object_model.columnCount()):
                    index = self.object_model.index(row, column)
                    da = self.object_model.itemData(index)
                    self.your_pandas_data_objects.iat[row, column] = da[0]
            book = load_workbook(self.testObjectsSelectedFile)
            writer = ExcelWriter(self.testObjectsSelectedFile, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            self.your_pandas_data_objects.to_excel(writer, self.Object_sheet,index=False)
            writer.save()

    def on_click_save_as_test_Objects(self):
        if self.testObjectsSelectedFile is None:
            Tk().withdraw()
            messagebox.showinfo("Error", "Please Select the Test Objects Sheet")
        elif self.testObjectsSelectedFile is not None:
            Tk().withdraw()
            FileName = asksaveasfilename(
                initialdir=os.path.dirname(os.getcwd()) + '/Selenium/TestData',
                filetypes=[("Excel files", "Script_*.xlsx .xls")],
                defaultextension=".xlsx",
                initialfile="Script_*")
            if FileName is not None:
                for row in range(self.object_model.rowCount()):
                    for column in range(self.object_model.columnCount()):
                        index = self.object_model.index(row, column)
                        da = self.object_model.itemData(index)
                        self.your_pandas_data_objects.iat[row, column] = da[0]
                writer = ExcelWriter(FileName)
                self.your_pandas_data_objects.to_excel(writer, self.Object_sheet,index=False)
                writer.save()
            else:
                pass
        else:
            pass

    def select_from_comboBox_TestObjects(self):
        selectedSheet=self.comboBox_TestObjects.currentText()
        if self.testObjectsSelectedFile:
            wb = load_workbook(self.testObjectsSelectedFile)
            item = self.verticalLayout_20.takeAt(0)
            w = item.widget()
            w.deleteLater()
            self.Object_sheet = selectedSheet
            self.object_view = QTableView()
            a = self.testObjectsSelectedFile
            df = pd.read_excel(a, sheet_name=self.Object_sheet)
            self.your_pandas_data_objects = df.replace(np.nan, '', regex=True)
            header = list(self.your_pandas_data_objects.columns.values)
            self.object_model = pandatest.PandasModel(self.your_pandas_data_objects, header)
            self.object_view.setModel(self.object_model)
            self.verticalLayout_20.addWidget(self.object_view)
        else:
            pass
    def on_click_open_driver_test_execution(self):
            dir_path = os.path.dirname(os.path.realpath(__file__))
            self.ExecutionFile =os.path.dirname(dir_path) + '/Selenium/TestData/Execution.xlsx'
            if self.ExecutionFile :
                item=self.verticalLayout_30.takeAt(0)
                print(item)
                w=item.widget()
                print(w)
                w.deleteLater()
                self.execution_view = QTableView()
                self.ExecutionFile

                df = pd.read_excel(self.ExecutionFile)

                self.your_pandas_data_execution = df.replace(np.nan, '', regex=True)
                header = list(self.your_pandas_data_execution.columns.values)
                self.execution_model = pandatest.PandasModel(self.your_pandas_data_execution, header)
                self.execution_view.setModel(self.execution_model)
                self.verticalLayout_30.addWidget(self.execution_view)
            else:
                    pass

    def on_click_execute_test_execution(self):
        hq=hqdriver.hqDriver
        hq.execute()

    def on_click_save_test_execution(self):
            if self.ExecutionFile is  None:
                    Tk().withdraw()
                    messagebox.showinfo("Error", "Please Select the Execution Sheet")
            else:
                for row in range(self.execution_model.rowCount()):
                    for column in range(self.execution_model.columnCount()):
                        index = self.execution_model.index(row, column)
                        da = self.execution_model.itemData(index)
                        self.your_pandas_data_execution.iat[row, column] = da[0]
                self.your_pandas_data_execution.to_excel(self.ExecutionFile,index=False)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "UI4AutomationFrameworks"))
        self.RecordPlayBtn.setText(_translate("MainWindow", "Record/Play"))
        self.label_17.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.RecordTab), _translate("MainWindow", "Record"))
        self.label_WorkBookName_ScriptGenerate.setText(_translate("MainWindow", "WorkBook Name :"))
        self.label_WorkBookName2_ScriptGenerate.setText(_translate("MainWindow", "<WorkBook Name>"))
        self.pushButton_Open_ScriptGenerate.setText(_translate("MainWindow", "Open"))
        self.pushButton_Save_ScriptGenerate.setText(_translate("MainWindow", "Save"))
        self.pushButton_GenerateScript_ScriptGenerate.setText(_translate("MainWindow", "Generate Script"))
        self.label_18.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.ScriptGenerateTab), _translate("MainWindow", "Script Generate"))
        self.label_WorkbookName_TestManagement.setText(_translate("MainWindow", "Workbook Name:"))
        self.label_WorkbookName1_TestManagement.setText(_translate("MainWindow", "<WorkBook Name>"))
        self.pushButton_Open_TestManagement.setText(_translate("MainWindow", "Open"))
        self.pushButton_Open_TestManagement_2.setText(_translate("MainWindow", "Save"))
        self.checkBoxCreateDefectsAutomatically.setText(_translate("MainWindow", "Create Defects Automatically"))
        self.label_ChoosetheTestManagementTool.setText(_translate("MainWindow", "Choose the Test Management Tool"))
        self.radioButton_ALM.setText(_translate("MainWindow", "ALM"))
        self.radioButton_JIRA.setText(_translate("MainWindow", "JIRA"))
        self.radioButton_NA.setText(_translate("MainWindow", "NA"))
        self.label_19.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestManagementTab), _translate("MainWindow", "Test Management"))
        self.label_MasterWorkBookName.setText(_translate("MainWindow", "Master WorkBook Name:"))
        self.label_MasterWorkBookName_2.setText(_translate("MainWindow", "<Master WorkBook Name>"))
        self.TestSetsOpenBtn_2.setText(_translate("MainWindow", "Open"))
        self.TestSetsSaveBtn_2.setText(_translate("MainWindow", "Save"))
        self.TestSetsSaveAsBtn.setText(_translate("MainWindow", "Save As"))
        self.TestSetsAddNewBtn.setText(_translate("MainWindow", "Add New"))
        self.TestSetsAddRowBtn.setText(_translate("MainWindow", "Add Row"))
        self.TestSetsDeleteRowBtn.setText(_translate("MainWindow", "Delete Row"))
        self.label_20.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestSetsTab), _translate("MainWindow", "Test Sets"))
        self.label_SelecttheScriptSheet.setText(_translate("MainWindow", "Select the Script Sheet"))
        self.pushButton_Open_TestScripts.setText(_translate("MainWindow", "Open"))
        self.pushButton_Save_TestScripts.setText(_translate("MainWindow", "Save"))
        self.pushButton_SaveAs_TestScripts.setText(_translate("MainWindow", "Save As"))
        self.pushButton_AddNew_TestScripts.setText(_translate("MainWindow", "Add New"))
        self.pushButton_AddRow_TestScripts.setText(_translate("MainWindow", "Add Row"))
        self.pushButton_deleteRow_TestScripts.setText(_translate("MainWindow", "Delete Row"))
        self.label_TestAutomationToolAndTechnology.setText(_translate("MainWindow", "Test Automation Tool & Technology"))
        self.comboBox_tool.setItemText(0, _translate("MainWindow", "Selenium"))
        self.comboBox_technology.setItemText(0, _translate("MainWindow", "Web"))
        self.label_21.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestScriptsTab), _translate("MainWindow", "Test Scripts"))
        self.label_SleecttheComponentSheet.setText(_translate("MainWindow", "Select the Component sheet"))
        self.pushButton_Open_TestComponents.setText(_translate("MainWindow", "Open"))
        self.pushButton_Save_TestComponents.setText(_translate("MainWindow", "Save"))
        self.pushButton_SaveAs_TestComponents.setText(_translate("MainWindow", "Save As"))
        self.pushButton_AddNew_TestComponents.setText(_translate("MainWindow", " Add New"))
        self.pushButton_AddRow_TestComponents.setText(_translate("MainWindow", "Add Row"))
        self.pushButton_DeleteRow_TestComponents.setText(_translate("MainWindow", "Delete Row"))
        self.label_22.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestComponentsTab), _translate("MainWindow", "Test Components"))
        self.label_SelecttheTestDataSheet.setText(_translate("MainWindow", "Select the Test Data Sheet"))
        self.pushButton_Open_TestData.setText(_translate("MainWindow", "Open"))
        self.pushButton_Save_TestData.setText(_translate("MainWindow", "Save"))
        self.pushButton_SaveAs_TestData.setText(_translate("MainWindow", "Save As"))
        self.pushButton_AddNew_TestData.setText(_translate("MainWindow", "Add New"))
        self.pushButton_AddRow_TestData.setText(_translate("MainWindow", "Add Row"))
        self.pushButton_DeleteRow_TestData.setText(_translate("MainWindow", "Delete Row"))
        self.pushButton_AddColumn_TestData.setText(_translate("MainWindow", "Add Column"))
        self.pushButton_DeleteColumn_TestData.setText(_translate("MainWindow", "Delete Column"))
        self.label_23.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestDataTab), _translate("MainWindow", "Test Data"))
        self.label_SelecttheObjectsExcel.setText(_translate("MainWindow", "Select the Objects Excel"))
        self.pushButton_Open_TestObjects.setText(_translate("MainWindow", "Open"))
        self.pushButton_Save_TestObjects.setText(_translate("MainWindow", "Save"))
        self.pushButton_SaveAs_TestObjects.setText(_translate("MainWindow", "Save As"))
        self.pushButton_AddNew_TestObjects.setText(_translate("MainWindow", "Add New"))
        self.pushButton_AddRow_TestObjects.setText(_translate("MainWindow", "Add Row"))
        self.pushButton_DeleteRow_TestObjects.setText(_translate("MainWindow", "Delete Row"))
        self.pushButton_AddColumn_TestObjects.setText(_translate("MainWindow", "Add Column"))
        self.label_24.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestObjectsTab), _translate("MainWindow", "Test Objects"))
        self.label_5.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.pushButton_OpenDriver_TestExecution.setText(_translate("MainWindow", "Open Driver"))
        self.pushButton_Save_TestExecution.setText(_translate("MainWindow", "Save"))
        self.label_TestExecutionTool.setText(_translate("MainWindow", "Test Execution Tool"))
        self.comboBox_TestExecution.setItemText(0, _translate("MainWindow", "Selenium"))
        self.pushButton_Execute_TestExecution.setText(_translate("MainWindow", "Execute"))
        self.pushButton_AddRow_TestExecution.setText(_translate("MainWindow", "Add Row"))
        self.pushButton_DeleteRow_TestExecution.setText(_translate("MainWindow", "Delete Row"))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestExecutionTab), _translate("MainWindow", "Test Execution"))
        self.label_TestResultPath.setText(_translate("MainWindow", "Test Result Path"))
        self.label_defectSummary.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt; font-weight:600;\">Defect Summary</span></p></body></html>"))
        self.label_ExecutionSummary.setText(_translate("MainWindow", "<html><head/><body><p><span style=\" font-size:12pt; font-weight:600;\">Execution Summary</span></p></body></html>"))
        self.label_25.setText(_translate("MainWindow", "@ Copyright IBM Corp.2017 All Rights Reserverd."))
        self.Maintab.setTabText(self.Maintab.indexOf(self.TestReportsTab), _translate("MainWindow", "Test Reports"))
        self.label_10.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:14pt; font-weight:600;\">UIAF - User Interface for Automation Frameworks V1.0</span></p><p>This UIAF launchpad consists of frameworks for the following tools, <br/>1. HP UFT<br/>2. IBM RFT<br/>3. Selenium<br/>4. Appium<br/>5. Record and Run Facility<br/>6. Script Generator<br/><br/><span style=\" text-decoration: underline;\">STeF for UFT:</span><br/>Scriptless Test Framework (STeF) is developed for HP UFT with integrations with tools like HP ALM and JIRA<br/>Contributed By:<br/>GTA Team<br/><br/><span style=\" text-decoration: underline;\">IBM RFT:</span><br/>Framework developed with Hybrid approach using Keywords and Data Parameterization. Reusable methods were developed and converted to keywords<br/>Contributed By:<br/>Framework - GTA Team (Rajkumar Natarajan)<br/>Reusable methods - TAC Team<br/><br/><span style=\" text-decoration: underline;\">Selenium:</span><br/>Framework developed with Hybrid approach using Keywords and Data Parameterization. Reusable methods were developed and converted to keywords<br/>Contributed By:<br/>Framework: GTA Team (Rajkumar Natarajan)<br/>GTA Team<br/><br/><span style=\" text-decoration: underline;\">SoapUI:</span><br/>This SoapUI framework assetization is in progress..<br/><br/><br/><span style=\" text-decoration: underline;\">Note:</span> UIAF launchpad is developed by GTA Team and property of IBM. For any modifications/changes reach out to,<br/>Rajkumar Natarajan(rnatara8@in.ibm.com)<br/>Margaret D(margaret.d@in.ibm.com)</p></body></html>"))
        self.Maintab.setTabText(self.Maintab.indexOf(self.AboutTab), _translate("MainWindow", "About"))
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuConfigure.setTitle(_translate("MainWindow", "Configure"))
        self.actionConfigure.setText(_translate("MainWindow", "SopaUI"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
